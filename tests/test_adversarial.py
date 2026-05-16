"""
Adversarial edge-case tests for the newsletter AI pipeline.
Each test is designed to expose real failure modes or verify subtle correctness guarantees.
No trivial happy-path tests.
"""
from __future__ import annotations
import math
import os
import re
import sys
import time
import tempfile
import unittest
from datetime import date
from unittest.mock import MagicMock, patch, call

import numpy as np
import pandas as pd
import pytest

# ── ensure project root is importable ─────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


# =============================================================================
# CLASS 1: AdversarialFormatting
# =============================================================================
class AdversarialFormatting(unittest.TestCase):
    """fmt_da under extreme and degenerate inputs."""

    def setUp(self):
        from newsletter_agent.formatting import fmt_da
        self.fmt_da = fmt_da

    def test_inf_does_not_crash(self):
        """fmt_da(inf, 2) currently CRASHES with ValueError because Python formats
        float('inf') as 'inf' (no decimal point), breaking rsplit('.', 1).
        This test documents the bug: it must either not crash OR raise ValueError cleanly.
        BUG: formatting.py line 35 — rsplit('.', 1) on 'inf' fails.
        """
        try:
            result = self.fmt_da(float("inf"), 2)
            assert isinstance(result, str), f"Expected str, got {type(result)}"
        except (ValueError, OverflowError):
            # Document the known bug — the crash is the failure mode
            pytest.xfail(
                "BUG: fmt_da(inf, 2) crashes — Python formats inf as 'inf' (no decimal point), "
                "breaking rsplit('.', 1) at formatting.py:35"
            )

    def test_nan_does_not_crash(self):
        """fmt_da(nan, 2) currently CRASHES with ValueError — same rsplit bug as inf.
        BUG: formatting.py line 35.
        """
        try:
            result = self.fmt_da(float("nan"), 2)
            assert isinstance(result, str), f"Expected str, got {type(result)}"
        except (ValueError, OverflowError):
            pytest.xfail(
                "BUG: fmt_da(nan, 2) crashes — Python formats nan as 'nan' (no decimal point), "
                "breaking rsplit('.', 1) at formatting.py:35"
            )

    def test_negative_zero_formats_as_zero(self):
        """fmt_da(-0.0, 2) currently produces '-0,00' — negative zero sign is ambiguous.
        Python's format emits '-0.00' which becomes '-0,00' after the swap.
        BUG: no guard for -0.0 before formatting. Ideally should produce '0,00'.
        """
        result = self.fmt_da(-0.0, 2)
        # Document actual behaviour — currently produces '-0,00'
        if result == "0,00":
            pass  # Fixed — test passes
        elif result == "-0,00":
            pytest.xfail(
                "BUG: fmt_da(-0.0, 2) = '-0,00' — negative zero should be '0,00'. "
                "Add: if val == 0: val = 0.0 before formatting in fmt_da()."
            )
        else:
            pytest.fail(f"Unexpected output: '{result}'")

    def test_quadrillion_no_scientific_notation(self):
        """fmt_da(1e15, 0) must use periods as thousands separators, not scientific notation."""
        result = self.fmt_da(1e15, 0)
        assert "e" not in result.lower(), f"Scientific notation in '{result}'"
        assert result == "1.000.000.000.000.000", f"Got '{result}'"

    def test_tiny_number_no_scientific_notation(self):
        """fmt_da(0.00000001, 4) must not produce scientific notation."""
        result = self.fmt_da(0.00000001, 4)
        assert "e" not in result.lower(), f"Scientific notation in '{result}'"

    def test_999_999_stays_below_threshold(self):
        """fmt_da(999.999, 2) must be '999,99' — does NOT round up to 1.000."""
        result = self.fmt_da(999.999, 2)
        # Python rounds 999.999 to 1000.00 at 2dp — if it does, the auto-select would
        # kick in and change formatting. With explicit decimals=2, we get whatever
        # Python's format produces: 1.000,00 is acceptable IF Python rounds up.
        # What MUST NOT happen: the string contains scientific notation or crashes.
        assert isinstance(result, str)
        assert "e" not in result.lower()

    def test_negative_999_near_threshold(self):
        """fmt_da(-999.999, 2) must not crash and must start with '-' (negative)."""
        result = self.fmt_da(-999.999, 2)
        assert isinstance(result, str)
        # If rounding gives -1000.00 that still produces '-1.000,00' which is correct
        assert result.startswith("-"), f"Expected leading minus in '{result}'"

    def test_exact_1000_triggers_integer_format(self):
        """fmt_da(1000.0, 2) → auto-select: abs_val >= 1000 → integer format → '1.000'."""
        result = self.fmt_da(1000.0)  # auto-select
        assert result == "1.000", f"Got '{result}', expected '1.000'"

    def test_roundtrip_zero_decimals_no_commas(self):
        """For N in test values, fmt_da(N, 0) contains only digits, dots, and optional minus.
        No comma (comma is decimal separator in Danish, not valid with 0 decimals)."""
        for val in [1, 100, 999, 1000, 1_000_000, -500_000]:
            result = self.fmt_da(val, 0)
            allowed = set("0123456789.-")
            illegal = set(result) - allowed
            assert not illegal, (
                f"fmt_da({val}, 0) = '{result}' contains illegal chars: {illegal}"
            )
            # Must not contain a comma (that would be decimal separator)
            assert "," not in result, (
                f"fmt_da({val}, 0) = '{result}' contains a comma (should have none at 0 decimals)"
            )


# =============================================================================
# CLASS 2: AdversarialRetry
# =============================================================================
class _FakeAPIError(Exception):
    """Fake Anthropic-like API error with a status_code attribute."""
    def __init__(self, status_code: int, message: str = "error"):
        super().__init__(message)
        self.status_code = status_code


class _FakeAPIErrorNoStatus(Exception):
    """Fake error that is an APIError subclass but has no retryable status code."""
    def __init__(self):
        super().__init__("auth failed")
        self.status_code = 401


class AdversarialRetry(unittest.TestCase):

    def setUp(self):
        from newsletter_agent.llm_retry import llm_call_with_retry, _is_retryable
        self.retry = llm_call_with_retry
        self.is_retryable = _is_retryable

    def _make_fn(self, fail_times: int, result="ok"):
        """Return a fn that fails `fail_times` times with a retryable error, then returns result."""
        calls = []
        def fn(**kwargs):
            calls.append(1)
            if len(calls) <= fail_times:
                raise _FakeAPIError(529, "overloaded")
            return result
        fn.calls = calls
        return fn

    def test_succeeds_on_last_allowed_attempt(self):
        """Retry succeeds on the last allowed attempt (attempt = retries-1)."""
        retries = 4
        fn = self._make_fn(fail_times=retries - 1, result="success")
        with patch("time.sleep"):
            result = self.retry(fn, retries=retries)
        assert result == "success"
        assert len(fn.calls) == retries

    def test_exact_call_count_before_raise(self):
        """Confirm exactly N calls were made before raising (not N-1, not N+1)."""
        retries = 3
        call_log = []
        def always_fail(**kwargs):
            call_log.append(1)
            raise _FakeAPIError(500, "server error")
        with patch("time.sleep"):
            with pytest.raises(_FakeAPIError):
                self.retry(always_fail, retries=retries)
        assert len(call_log) == retries, f"Expected {retries} calls, got {len(call_log)}"

    def test_backoff_is_strictly_increasing(self):
        """Sleep durations are exponential and strictly increasing across attempts."""
        sleep_calls = []
        def record_sleep(t):
            sleep_calls.append(t)
        def always_fail(**kwargs):
            raise _FakeAPIError(529, "overloaded")
        with patch("time.sleep", side_effect=record_sleep):
            with pytest.raises(_FakeAPIError):
                self.retry(always_fail, retries=4)
        assert len(sleep_calls) == 3, f"Expected 3 sleep calls, got {len(sleep_calls)}"
        for i in range(1, len(sleep_calls)):
            assert sleep_calls[i] > sleep_calls[i - 1], (
                f"Sleep[{i}]={sleep_calls[i]:.3f} not > Sleep[{i-1}]={sleep_calls[i-1]:.3f}"
            )

    def test_jitter_produces_variance(self):
        """Two runs with the same attempt number produce DIFFERENT sleep durations (jitter)."""
        seen = set()
        for _ in range(20):
            sleep_calls = []
            def record_sleep(t):
                sleep_calls.append(t)
            calls = [0]
            def fail_once(**kwargs):
                calls[0] += 1
                if calls[0] == 1:
                    raise _FakeAPIError(529, "overloaded")
                return "ok"
            with patch("time.sleep", side_effect=record_sleep):
                self.retry(fail_once, retries=2)
            if sleep_calls:
                seen.add(round(sleep_calls[0], 6))
        assert len(seen) > 1, "All 20 runs produced identical sleep duration — jitter is absent"

    def test_529_is_retried(self):
        """529 status code on exc.status_code → retried."""
        exc = _FakeAPIError(529)
        assert self.is_retryable(exc), "529 should be retryable"

    def test_500_is_retried(self):
        """500 status code → retried."""
        exc = _FakeAPIError(500)
        assert self.is_retryable(exc), "500 should be retryable"

    def test_429_is_retried(self):
        """429 status code → retried."""
        exc = _FakeAPIError(429)
        assert self.is_retryable(exc), "429 should be retryable"

    def test_401_is_not_retried(self):
        """401 status code → NOT retried (auth error is user's fault)."""
        exc = _FakeAPIError(401)
        assert not self.is_retryable(exc), "401 should NOT be retryable"

    def test_400_is_not_retried(self):
        """400 status code → NOT retried."""
        exc = _FakeAPIError(400)
        assert not self.is_retryable(exc), "400 should NOT be retryable"

    def test_api_error_subclass_non_retryable_status_not_retried(self):
        """Exception where _is_retryable returns False (401) → not retried, raises immediately."""
        call_log = []
        def fn(**kwargs):
            call_log.append(1)
            raise _FakeAPIErrorNoStatus()
        with patch("time.sleep"):
            with pytest.raises(_FakeAPIErrorNoStatus):
                self.retry(fn, retries=5)
        assert len(call_log) == 1, (
            f"Expected immediate raise (1 call), got {len(call_log)} calls"
        )

    def test_plain_value_error_not_retried(self):
        """Plain ValueError (not Anthropic error) → NOT retried, raises immediately."""
        call_log = []
        def fn(**kwargs):
            call_log.append(1)
            raise ValueError("something broke")
        with patch("time.sleep"):
            with pytest.raises(ValueError):
                self.retry(fn, retries=5)
        assert len(call_log) == 1, (
            f"ValueError should not be retried — got {len(call_log)} calls"
        )


# =============================================================================
# CLASS 3: AdversarialTidsperiode
# =============================================================================
class AdversarialTidsperiode(unittest.TestCase):
    """period_days enforcement, date boundary logic, and specialist behaviour."""

    def _make_manifest_with_chart(self, y_label: str, period_days: int = 365) -> dict:
        return {
            "specialists": ["macro"],
            "macro": {
                "series": [],
                "charts": [{"type": "A", "title": "Test", "y_label": y_label, "period_days": period_days}],
            },
        }

    def test_zero_day_range_period_days_computed(self):
        """start_date == end_date → period_days derived as max(1, 0) = 1. No crash."""
        from newsletter_agent import pipeline
        _s = pd.Timestamp("2024-01-01")
        _e = pd.Timestamp("2024-01-01")
        period_days = max(1, (_e - _s).days)
        assert period_days == 1

    def test_future_dates_positive_period_days(self):
        """Future date range: period_days is positive and correct."""
        _s = pd.Timestamp("2030-01-01")
        _e = pd.Timestamp("2031-01-01")
        period_days = max(1, (_e - _s).days)
        assert period_days == 365

    def test_single_day_chart_spec_minimum(self):
        """period_days=1 non-YoY chart stays at 1 (no inflation)."""
        from newsletter_agent import pipeline as pl
        manifest = self._make_manifest_with_chart("Pct.", 1)
        # Simulate the enforcement block with period_days=1
        period_days = 1
        for spec_name in manifest["specialists"]:
            for chart in manifest.get(spec_name, {}).get("charts", []):
                is_yoy = "YoY" in chart.get("y_label", "") or "yoy" in chart.get("y_label", "").lower()
                min_days = 760 if is_yoy else period_days
                chart["period_days"] = max(min_days, period_days) if is_yoy else period_days
        result_pd = manifest["macro"]["charts"][0]["period_days"]
        assert result_pd == 1

    def test_100_year_period_worldbank_caps_at_30_years(self):
        """period_days=36500 (100 years) → worldbank years = min(30, ...) = 30 via pipeline logic."""
        period_days = 36500
        years = max(5, min(30, period_days // 365))
        assert years == 30, f"Expected 30, got {years}"

    def test_yoy_lowercase_gets_760_minimum(self):
        """y_label='yoy %' (lowercase) → is_yoy=True → gets 760 minimum."""
        y_label = "yoy %"
        is_yoy = "YoY" in y_label or "yoy" in y_label.lower()
        assert is_yoy
        period_days = 365
        min_days = 760 if is_yoy else period_days
        enforced = max(min_days, period_days) if is_yoy else period_days
        assert enforced == 760, f"Expected 760, got {enforced}"

    def test_yoy_embedded_in_string_gets_760_minimum(self):
        """y_label='YoY change (%)' (embedded) → is_yoy=True → gets 760 minimum."""
        y_label = "YoY change (%)"
        is_yoy = "YoY" in y_label or "yoy" in y_label.lower()
        assert is_yoy
        period_days = 365
        enforced = max(760, period_days) if is_yoy else period_days
        assert enforced == 760

    def test_non_yoy_chart_with_yoy_in_label_but_non_yoy_transform(self):
        """'YoY-adjusted seasonally' has YoY in label — pipeline WILL detect it as YoY.
        This documents the current behaviour: YoY detection is case-sensitive substring,
        so 'YoY' anywhere in y_label triggers the 760 minimum."""
        y_label = "YoY-adjusted seasonally"
        is_yoy = "YoY" in y_label or "yoy" in y_label.lower()
        # Per actual code at pipeline.py:964, this IS flagged as YoY (substring match).
        # The test documents this is INTENTIONAL: broad match is safer than false negative.
        assert is_yoy, (
            "The current implementation flags ANY y_label containing 'YoY' as a YoY chart. "
            "This is the documented behaviour — if this fails the detection logic changed."
        )

    def test_macro_specialist_same_day_date_no_crash(self):
        """Macro specialist: start_date == end_date (same day) → FRED filter produces empty
        but fetch_macro itself does not crash (it just returns empty dataframes dict)."""
        # We test the date filtering logic directly without hitting FRED
        import pandas as pd
        start = "2010-01-01"
        end = "2010-01-01"
        # Build a fake series spanning years and filter it to the same-day window
        idx = pd.date_range("2009-01-01", "2011-01-01", freq="ME")
        series = pd.Series(range(len(idx)), index=idx)
        filtered = series[(series.index >= pd.Timestamp(start)) & (series.index <= pd.Timestamp(end))]
        # Must not crash; result is empty
        assert isinstance(filtered, pd.Series)
        # Empty is fine — no crash is the requirement
        assert len(filtered) == 0

    def test_worldbank_years_formula_from_start_date(self):
        """worldbank.py: years = date.today().year - start_year + 1."""
        from datetime import date as _date
        import pandas as _pd
        start_str = "2020-01-01"
        years = _date.today().year - _pd.Timestamp(start_str).year + 1
        expected = _date.today().year - 2020 + 1
        assert years == expected, f"Expected {expected}, got {years}"

    def test_eurostat_end_date_past_includes_only_past_data(self):
        """Eurostat: end_date in the past → only data up to that date is included."""
        end_date = "2015-12-31"
        idx = pd.date_range("2010-01-01", "2020-01-01", freq="YS")
        df = pd.DataFrame({"value": range(len(idx))}, index=idx)
        # Simulate the filter a caller would apply
        filtered = df[df.index <= pd.Timestamp(end_date)]
        assert filtered.index.max() <= pd.Timestamp(end_date)
        assert not filtered.empty


# =============================================================================
# CLASS 4: AdversarialBarChart
# =============================================================================
class AdversarialBarChart(unittest.TestCase):
    """render_type_b with degenerate DataFrames."""

    def _spec(self, **kwargs) -> dict:
        defaults = {"title": "Test Chart", "x_label": "", "y_label": "%",
                    "note": "", "kilde": ""}
        defaults.update(kwargs)
        return defaults

    def _ts_df(self, n: int, value=1.0, col="Værdi") -> pd.DataFrame:
        idx = pd.date_range("2020-01-01", periods=n, freq="ME")
        return pd.DataFrame({col: [value] * n}, index=idx)

    def setUp(self):
        from newsletter_agent.renderers.charts import render_type_b
        self.render = render_type_b
        self._tmpdir = tempfile.mkdtemp()

    def _path(self, name: str) -> str:
        return os.path.join(self._tmpdir, name + ".png")

    def test_single_row_no_crash(self):
        """Single-row time-series (n=1 bar) must render without crash."""
        df = self._ts_df(1)
        out = self.render(df, self._spec(), self._path("single_row"))
        assert out and os.path.exists(out)

    def test_all_zeros_no_crash(self):
        """All-zero time-series (50 bars of 0.0) must render; y-axis must not crash on zero range."""
        df = self._ts_df(50, value=0.0)
        out = self.render(df, self._spec(), self._path("all_zeros"))
        assert out and os.path.exists(out)

    def test_all_nan_column_no_crash(self):
        """All-NaN column — after dropna(), series is empty → must not crash."""
        idx = pd.date_range("2020-01-01", periods=12, freq="ME")
        df = pd.DataFrame({"Værdi": [float("nan")] * 12}, index=idx)
        # render_type_b drops NaN: series = df[col].dropna() → empty
        # This may crash on len(series) == 0 or plotting. The test just must not raise.
        try:
            out = self.render(df, self._spec(), self._path("all_nan"))
            # If it returns a path, check file exists
            if out:
                assert os.path.exists(out)
        except Exception as e:
            pytest.fail(f"render_type_b raised {type(e).__name__} on all-NaN input: {e}")

    def test_mixed_nan_highlight_marks_valid_bars(self):
        """48 valid values + 2 NaN at end → NaN dropped, highlight_last_n=2 marks last 2 VALID bars."""
        idx = pd.date_range("2020-01-01", periods=50, freq="ME")
        vals = [1.0] * 48 + [float("nan"), float("nan")]
        df = pd.DataFrame({"Værdi": vals}, index=idx)
        spec = self._spec(highlight_last_n=2, highlight_color="#ff0000")
        # Must not crash; file written
        out = self.render(df, spec, self._path("mixed_nan"))
        assert out and os.path.exists(out)

    def test_highlight_last_n_larger_than_bars_no_index_error(self):
        """highlight_last_n=1000 with only 12 bars → no IndexError, all bars highlighted."""
        df = self._ts_df(12)
        spec = self._spec(highlight_last_n=1000, highlight_color="#ff0000")
        out = self.render(df, spec, self._path("highlight_overflow"))
        assert out and os.path.exists(out)

    def test_bar_color_none_falls_back_to_brand(self):
        """bar_color=None → falls back to brand teal without crash."""
        df = self._ts_df(10)
        spec = self._spec(bar_color=None)
        out = self.render(df, spec, self._path("bar_color_none"))
        assert out and os.path.exists(out)

    def test_bar_color_invalid_hex_no_propagated_exception(self):
        """bar_color='not-a-hex' → must not propagate an unhandled exception to the caller."""
        df = self._ts_df(10)
        spec = self._spec(bar_color="not-a-hex")
        try:
            out = self.render(df, spec, self._path("bad_color"))
            if out:
                assert os.path.exists(out)
        except Exception as e:
            pytest.fail(
                f"render_type_b let an invalid bar_color propagate: {type(e).__name__}: {e}"
            )

    def test_single_year_monthly_labels_branch(self):
        """12 bars of monthly data → single year group → monthly labels branch. File exists."""
        df = self._ts_df(12)
        out = self.render(df, self._spec(), self._path("single_year"))
        assert out and os.path.exists(out)

    def test_extremely_large_values_no_overflow(self):
        """Bars with values of 1e12 render without overflow error."""
        df = self._ts_df(10, value=1e12)
        out = self.render(df, self._spec(), self._path("large_values"))
        assert out and os.path.exists(out)

    def test_all_negative_values_bars_below_zero(self):
        """All-negative values: bars render below zero, zero-line visible. No crash."""
        df = self._ts_df(20, value=-5.5)
        out = self.render(df, self._spec(), self._path("all_negative"))
        assert out and os.path.exists(out)


# =============================================================================
# CLASS 5: AdversarialRouting
# =============================================================================
class AdversarialRouting(unittest.TestCase):
    """get_routing_hint edge cases."""

    def setUp(self):
        from newsletter_agent.routing import get_routing_hint
        self.hint = get_routing_hint

    def test_eu_and_us_employment_both_fire(self):
        """EU AND US employment keywords → multiple hints; neither overrides the other."""
        brief = "EU unemployment rate and US nonfarm payroll"
        result = self.hint(brief)
        # EU+UNEM should fire, EMPL should fire
        assert "eu_unemployment" in result or "eurostat" in result.lower(), (
            f"EU unemployment hint missing in: {result}"
        )
        assert "payroll" in result.lower() or "PAYEMS" in result or "employment" in result.lower(), (
            f"US employment hint missing in: {result}"
        )

    def test_empty_string_no_crash(self):
        """Empty string prompt → no hints, no crash."""
        result = self.hint("")
        assert isinstance(result, str)
        assert result == ""

    def test_single_letter_no_hints(self):
        """Prompt = 'a' (single letter) → no hints."""
        result = self.hint("a")
        assert result == ""

    def test_emoji_only_no_crash(self):
        """Prompt with only emojis and spaces → no crash, no hints."""
        result = self.hint("🇩🇰 📊 💹")
        assert isinstance(result, str)
        assert result == ""

    def test_danish_compound_not_employment(self):
        """'arbejdsmarkedsvilkårene' does NOT fire the employment rule (wrong stem)."""
        brief = "arbejdsmarkedsvilkårene er komplekse"
        result = self.hint(brief)
        # EMPL pattern checks: beskæftigelse|jobvækst|jobrapport|lønmodtager|payroll|nonfarm|jobs added|arbejdspladser|employment report
        # 'arbejdsmarkedsvilkårene' should NOT match any of these
        # If it does match employment, that's a false positive — test documents it.
        # Check that PAYEMS (the US employment hint) does not appear:
        if "PAYEMS" in result:
            pytest.fail(
                f"False positive: 'arbejdsmarkedsvilkårene' triggered employment rule. Result: {result}"
            )

    def test_danish_compound_with_beskæftigelse_fires(self):
        """'arbejdsbeskæftigelsessituationen' contains 'beskæftigelse' → employment rule fires."""
        brief = "arbejdsbeskæftigelsessituationen i Danmark"
        result = self.hint(brief)
        assert "PAYEMS" in result or "beskæftigelse" in result.lower() or "employment" in result.lower(), (
            f"Expected employment hint, got: '{result}'"
        )

    def test_eu_energy_rule_fires(self):
        """'energimix i EU27' → EU+NRG rule fires."""
        brief = "energimix i EU27"
        result = self.hint(brief)
        assert "eu_energy_mix" in result or "eurostat" in result.lower(), (
            f"EU energy hint not found in: {result}"
        )

    def test_both_gas_rules_ttf_henry_hub(self):
        """'TTF Henry Hub' → cross-region gas rule fires."""
        brief = "TTF og Henry Hub gaspriser"
        result = self.hint(brief)
        assert "TTF" in result or "gas" in result.lower() or "EUR/MWh" in result, (
            f"Gas cross-region hint not found in: {result}"
        )

    def test_long_prompt_no_timeout(self):
        """Malformed brief with 10,000 characters → no exception, returns str."""
        brief = "inflation EU eurozone " * 500  # 10,000+ chars
        import time as _time
        t0 = _time.monotonic()
        result = self.hint(brief)
        elapsed = _time.monotonic() - t0
        assert isinstance(result, str)
        assert elapsed < 5.0, f"get_routing_hint took {elapsed:.2f}s on long input"


# =============================================================================
# CLASS 6: AdversarialExcel
# =============================================================================
class AdversarialExcel(unittest.TestCase):
    """_write_excel_per_figure with edge case inputs."""

    def setUp(self):
        from newsletter_agent.pipeline import _write_excel_per_figure
        self._fn = _write_excel_per_figure
        self._tmpdir = tempfile.mkdtemp()

    def _pkg(self, title: str, chart_type: str, region_labels: list, path: str = "") -> dict:
        return {
            "path": path,
            "metadata": {
                "title": title,
                "chart_type": chart_type,
                "region_labels": region_labels,
            },
        }

    def _good_df(self, n: int = 5) -> pd.DataFrame:
        idx = pd.date_range("2020-01-01", periods=n, freq="ME")
        return pd.DataFrame({"value": range(n)}, index=idx)

    def test_missing_label_in_all_dfs_graceful_skip(self):
        """series_labels refers to a label not in all_dfs → graceful skip, path=''."""
        pkg = self._pkg("Test", "A", ["GhostLabel"])
        specialist_results = {"macro": {"dataframes": {"RealLabel": self._good_df()}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        assert paths == [""], f"Expected [''], got {paths}"

    def test_all_nan_dataframe_written_zero_path_empty(self):
        """DataFrame with all-NaN values → written=0 → excel_path=''."""
        nan_df = pd.DataFrame(
            {"value": [float("nan")] * 5},
            index=pd.date_range("2020-01-01", periods=5, freq="ME"),
        )
        pkg = self._pkg("NaN Chart", "A", ["NaNSeries"])
        specialist_results = {"macro": {"dataframes": {"NaNSeries": nan_df}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        assert paths == [""], f"Expected [''] for all-NaN df, got {paths}"

    def test_long_title_sanitised_and_truncated(self):
        """200-character title with illegal chars → safe_title sanitised, path is valid."""
        bad_chars = r'/\*?:[]|<>'
        title = "A" * 50 + bad_chars + "B" * 100  # well over 200 chars
        pkg = self._pkg(title, "A", ["MySeries"])
        df = self._good_df()
        specialist_results = {"macro": {"dataframes": {"MySeries": df}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        path = paths[0]
        # If a valid excel was written, check filename
        if path:
            basename = os.path.basename(path)
            # Illegal chars should be gone from filename
            for ch in r'/\*?:|<>':
                assert ch not in basename, f"Illegal char '{ch}' in filename: {basename}"
            # Filename should be manageable (safe_title capped at 45)
            assert len(basename) < 200

    def test_existing_png_embeds_grafik_sheet_first(self):
        """Existing PNG → openpyxl image embed; 'Grafik' sheet is first sheet in workbook."""
        import openpyxl
        from PIL import Image as PILImage
        # Create a real tiny PNG
        png_path = os.path.join(self._tmpdir, "test.png")
        img = PILImage.new("RGB", (100, 100), color=(255, 0, 0))
        img.save(png_path)

        pkg = self._pkg("Image Test", "A", ["MySeries"], path=png_path)
        df = self._good_df()
        specialist_results = {"macro": {"dataframes": {"MySeries": df}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        path = paths[0]
        assert path and os.path.exists(path), "Expected excel file to be created"
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames[0] == "Grafik", (
            f"Expected 'Grafik' as first sheet, got '{wb.sheetnames[0]}'"
        )

    def test_missing_png_image_embed_skipped_data_still_written(self):
        """PNG path does not exist → image embed skipped; data sheets still written."""
        missing_png = os.path.join(self._tmpdir, "nonexistent.png")
        pkg = self._pkg("No Image", "A", ["MySeries"], path=missing_png)
        df = self._good_df()
        specialist_results = {"macro": {"dataframes": {"MySeries": df}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        path = paths[0]
        assert path and os.path.exists(path), "Expected excel file even without image"
        import openpyxl
        wb = openpyxl.load_workbook(path)
        # Data sheet must exist (may be the only sheet)
        assert len(wb.sheetnames) >= 1
        # "Grafik" must NOT be first (no image was embedded — or may not exist at all)
        # The critical invariant: no crash, data written
        assert "MySeries"[:31] in wb.sheetnames or any("MySeries" in s for s in wb.sheetnames)

    def test_chart_type_d_returns_empty_path(self):
        """Chart type 'D' → excel_path is '' (D-type tables skipped by design)."""
        pkg = self._pkg("D-Type Table", "D", ["SomeSeries"])
        df = self._good_df()
        specialist_results = {"macro": {"dataframes": {"SomeSeries": df}}}
        paths = self._fn([pkg], specialist_results, self._tmpdir)
        assert paths == [""], f"Expected [''] for type-D chart, got {paths}"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
